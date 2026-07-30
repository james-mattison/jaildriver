#!/usr/bin/env python3
"""Archive every available SLO Sheriff booking-log date as normalized JSON.

For each date currently offered by the booking-log form, this program:

1. Selects that date.
2. Selects 0:00 in the Time dropdown.
3. Selects All Agencies in the Agency dropdown.
4. Clicks Get Log.
5. Parses the booking table with BeautifulSoup.
6. Expands HTML rowspan/colspan cells and groups all charges under one booking.
7. Writes booking_logs/YYYY-MM-DD.json as a bare JSON array.

The output format is:

[
  {
    "booking_id": "A00833107.1",
    "id": "710979",
    "lastname": "ALVAREZ",
    "firstname": "JAYDON",
    "middle": "DAYNE",
    "sex": "M",
    "dob": "2003-10-28",
    "booked_at": "2026-07-14 04:11:54",
    "charges": [
      {"code": "23103(A)", "name": "reckless driving"}
    ]
  }
]
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Sequence

import requests
from bs4 import BeautifulSoup, Tag
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support.ui import Select, WebDriverWait

SCRIPT_VERSION = "4.0-final-charge-normalizer"
START_URL = "https://www.slosheriff.org/WebLogs/BookingLog/"
DOJ_WORKBOOK_URL = "https://oag.ca.gov/system/files/media/CIBRS%20Offense%20Listing.xlsx"

EXPECTED_HEADERS = {
    "booking_number",
    "id_number",
    "name",
    "sex",
    "dob",
    "date_time",
    "charge",
}

HEADER_ALIASES = {
    "booking_number": "booking_number",
    "booking_no": "booking_number",
    "booking": "booking_number",
    "id_number": "id_number",
    "id_no": "id_number",
    "id": "id_number",
    "name": "name",
    "sex": "sex",
    "dob": "dob",
    "date_of_birth": "dob",
    "date_time": "date_time",
    "datetime": "date_time",
    "booking_date_time": "date_time",
    "booked_at": "date_time",
    "f_m_i": "f_m_i",
    "fmi": "f_m_i",
    "charge": "charge",
    "charges": "charge",
    "arrest_location": "arrest_location",
    "agency": "agency",
    "next_court_date": "next_court_date",
}

# Human-friendly names requested by the user.  The DOJ workbook remains the
# fallback for the thousands of sections not listed here.
FRIENDLY_OVERRIDES = {
    "23103(A)": "reckless driving",
    "21310": "concealed weapon (knife)",
    "485": "petty theft",
    "148(A)(1)": "resisting or obstructing a peace officer",
    "23152(A)": "driving under the influence",
    "23152(B)": "driving with a blood alcohol level of 0.08% or higher",
    "647(F)": "public intoxication",
    "11350(A)": "possession of a controlled substance",
    "11364(A)": "possession of drug paraphernalia",
    "11377(A)": "possession of a controlled substance",
    "11550(A)": "being under the influence of a controlled substance",
    "459": "burglary",
    "484(A)": "petty theft",
    "487(A)": "grand theft",
    "594(B)(1)": "felony vandalism",
    "594(B)(2)(A)": "misdemeanor vandalism",
    "602": "trespassing",
    "69": "resisting an executive officer",
    "422(A)": "criminal threats",
    "273.5(A)": "domestic violence",
    "243(E)(1)": "domestic battery",
    "25400(A)(1)": "carrying a concealed firearm",
    "25850(A)": "carrying a loaded firearm in public",
    "530.5(A)": "identity theft",
    "530.5(C)(1)": "possession of another person's identifying information",
}

CODE_TYPE_PRIORITY = {
    "PC": 0,
    "VC": 1,
    "HS": 2,
    "WI": 3,
    "BP": 4,
    "FG": 5,
    "US": 6,
    "ZZ": 99,
}


@dataclass(frozen=True)
class DateChoice:
    text: str
    value: str
    iso_date: str


@dataclass(frozen=True)
class OffenseCandidate:
    code_type: str
    section: str
    description: str
    offense_level: str
    active: bool


class OffenseTranslator:
    def __init__(self, workbook_path: Path | None):
        self.by_section: dict[str, list[OffenseCandidate]] = {}
        if workbook_path is not None and workbook_path.exists():
            self._load_workbook(workbook_path)

    def _load_workbook(self, path: Path) -> None:
        logging.info("Loading offense descriptions from %s", path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook["Offense Codes"]
            rows = sheet.iter_rows(values_only=True)
            headers = [normalize_header(str(value or "")) for value in next(rows)]
            indexes = {name: index for index, name in enumerate(headers)}

            required = {"code_type", "code_section", "offense_description"}
            missing = required - indexes.keys()
            if missing:
                raise RuntimeError(
                    f"DOJ workbook is missing required columns: {sorted(missing)}"
                )

            for row in rows:
                section = normalize_charge_code(cell_value(row, indexes["code_section"]))
                description = normalize_space(
                    str(cell_value(row, indexes["offense_description"]) or "")
                )
                if not section or not description:
                    continue

                code_type = normalize_space(
                    str(cell_value(row, indexes["code_type"]) or "")
                ).upper()
                offense_level = normalize_space(
                    str(cell_value(row, indexes.get("offense_level")) or "")
                ).upper()
                repeal_value = cell_value(row, indexes.get("repeal_date"))
                active = is_active_repeal_date(repeal_value)

                candidate = OffenseCandidate(
                    code_type=code_type,
                    section=section,
                    description=description,
                    offense_level=offense_level,
                    active=active,
                )
                self.by_section.setdefault(section, []).append(candidate)
        finally:
            workbook.close()

    def translate(self, raw_code: str) -> dict[str, str]:
        code_type_hint, code = split_code_type(raw_code)
        if not code:
            return {"code": "", "name": "unknown offense"}

        if code in FRIENDLY_OVERRIDES:
            return {"code": code, "name": FRIENDLY_OVERRIDES[code]}

        candidates = self.by_section.get(code, [])
        if code_type_hint:
            typed = [item for item in candidates if item.code_type == code_type_hint]
            if typed:
                candidates = typed

        if not candidates:
            return {"code": code, "name": "unknown offense"}

        selected = min(candidates, key=offense_candidate_score)
        return {
            "code": code,
            "name": friendly_offense_name(selected.description),
        }


def cell_value(row: Sequence[object], index: int | None) -> object | None:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def is_active_repeal_date(value: object | None) -> bool:
    if value is None:
        return True
    text = re.sub(r"\D", "", str(value))
    return not text or text.startswith("9999")


def offense_candidate_score(candidate: OffenseCandidate) -> tuple[int, int, int, int]:
    description = candidate.description.upper()
    qualification_penalty = sum(
        marker in description
        for marker in (
            "W/PRIOR",
            "WITH PRIOR",
            "PRIOR CONVICT",
            "HABITUAL",
            "ENHANCEMENT",
            "ATTEMPT",
        )
    )
    level_penalty = 0 if candidate.offense_level == "M" else 1
    return (
        0 if candidate.active else 1,
        qualification_penalty,
        CODE_TYPE_PRIORITY.get(candidate.code_type, 50),
        level_penalty,
    )


def friendly_offense_name(description: str) -> str:
    text = normalize_space(description).lower()
    replacements = {
        "poss ": "possession of ",
        "possess ": "possession of ",
        "ctrl subst": "controlled substance",
        "controlled subst": "controlled substance",
        "paraphernalia": "drug paraphernalia",
        "dui alcohol/0.08 percent": "driving with a blood alcohol level of 0.08% or higher",
        "carry concealed dirk or dagger": "concealed weapon (knife)",
        "appropriate lost property": "petty theft",
        "reckless driving:highway": "reckless driving",
        "obstruct/resist/etc public/peace officer/emergency med tech": (
            "resisting or obstructing a peace officer"
        ),
    }
    if text in replacements:
        return replacements[text]

    text = text.replace("w/", "with ")
    text = text.replace("/etc", "")
    text = text.replace("/", " or ")
    text = text.replace(":", " — ")
    text = re.sub(r"\betc\b", "", text)
    return normalize_space(text)


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def normalize_header(value: str) -> str:
    value = normalize_space(value).lower()
    value = value.replace("#", " number ")
    value = value.replace("date/time", "date time")
    value = re.sub(r"[^a-z0-9]+", "_", value).strip("_")
    return HEADER_ALIASES.get(value, value)


def normalize_charge_code(value: object) -> str:
    text = normalize_space(str(value or "")).upper()
    text = text.replace("§", "")
    text = re.sub(
        r"^(?:CA\s+)?(?:PENAL|VEHICLE|HEALTH\s*(?:AND|&)\s*SAFETY)\s+CODE\s*",
        "",
        text,
    )
    text = re.sub(r"^(PC|PEN|VC|VEH|HS|H&S|WI|W&I|BP|B&P|FG|F&G)\s*[:\-]?\s*", "", text)
    text = re.sub(r"\s*(PC|PEN|VC|VEH|HS|H&S|WI|W&I|BP|B&P|FG|F&G)$", "", text)
    text = re.sub(r"\s+", "", text)
    text = text.rstrip(".,;:")
    return text


def split_code_type(raw_value: str) -> tuple[str | None, str]:
    text = normalize_space(raw_value).upper()
    aliases = {
        "PEN": "PC",
        "VEH": "VC",
        "H&S": "HS",
        "W&I": "WI",
        "B&P": "BP",
        "F&G": "FG",
    }

    prefix = re.match(r"^(PC|PEN|VC|VEH|HS|H&S|WI|W&I|BP|B&P|FG|F&G)\b", text)
    suffix = re.search(r"\b(PC|PEN|VC|VEH|HS|H&S|WI|W&I|BP|B&P|FG|F&G)$", text)
    hint = None
    if prefix:
        hint = aliases.get(prefix.group(1), prefix.group(1))
    elif suffix:
        hint = aliases.get(suffix.group(1), suffix.group(1))
    return hint, normalize_charge_code(text)


def build_driver(headless: bool) -> webdriver.Chrome:
    options = webdriver.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--lang=en-US")
    options.add_argument(
        "--user-agent=SLOBookingLogArchiver/2.0 "
        "(public-record research; polite automated retrieval)"
    )
    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(60)
    return driver


def visible_selects(driver: WebDriver):
    return [
        element
        for element in driver.find_elements(By.TAG_NAME, "select")
        if element.is_displayed()
    ]


def find_select(driver: WebDriver, field_name: str):
    wanted = field_name.casefold()
    selects = visible_selects(driver)

    for element in selects:
        attributes = " ".join(
            filter(
                None,
                (
                    element.get_attribute("id"),
                    element.get_attribute("name"),
                    element.get_attribute("aria-label"),
                    element.get_attribute("title"),
                ),
            )
        ).casefold()
        if wanted in attributes:
            return element

    for label in driver.find_elements(By.TAG_NAME, "label"):
        if wanted not in normalize_space(label.text).casefold():
            continue
        target_id = label.get_attribute("for")
        if target_id:
            try:
                target = driver.find_element(By.ID, target_id)
                if target.tag_name.casefold() == "select" and target.is_displayed():
                    return target
            except NoSuchElementException:
                pass
        try:
            target = label.find_element(By.XPATH, ".//following::select[1]")
            if target.is_displayed():
                return target
        except NoSuchElementException:
            pass

    for element in selects:
        try:
            surrounding = normalize_space(
                element.find_element(By.XPATH, "./parent::*").text
            ).casefold()
        except (NoSuchElementException, StaleElementReferenceException):
            continue
        if wanted in surrounding:
            return element

    raise NoSuchElementException(
        f"Could not find visible {field_name!r} dropdown; "
        f"visible dropdown count={len(selects)}"
    )


def selectable_options(driver: WebDriver, field_name: str) -> list[tuple[str, str]]:
    result: list[tuple[str, str]] = []
    for option in Select(find_select(driver, field_name)).options:
        text = normalize_space(option.text)
        value = normalize_space(option.get_attribute("value") or "")
        if not option.is_enabled():
            continue
        if not text or text.casefold().startswith(("select ", "choose ", "please ")):
            continue
        if value.casefold() in {"", "-1", "none", "null"}:
            continue
        result.append((text, value))
    return result


def parse_date_text(text: str) -> str | None:
    cleaned = normalize_space(text)
    for pattern in (
        "%B %d, %Y",
        "%B %e, %Y",
        "%b %d, %Y",
        "%m/%d/%Y",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(cleaned, pattern).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def wait_for_stable_dates(driver: WebDriver, timeout: int) -> list[DateChoice]:
    deadline = time.monotonic() + timeout
    previous_signature: tuple[tuple[str, str], ...] | None = None
    stable_since: float | None = None
    latest: list[tuple[str, str]] = []

    while time.monotonic() < deadline:
        try:
            latest = selectable_options(driver, "date")
        except (NoSuchElementException, StaleElementReferenceException):
            latest = []

        signature = tuple(latest)
        if signature and signature == previous_signature:
            if stable_since is None:
                stable_since = time.monotonic()
            elif time.monotonic() - stable_since >= 1.5:
                choices = [
                    DateChoice(text=text, value=value, iso_date=iso)
                    for text, value in latest
                    if (iso := parse_date_text(text)) is not None
                ]
                if choices:
                    return choices
        else:
            previous_signature = signature
            stable_since = time.monotonic() if signature else None

        time.sleep(0.25)

    raise TimeoutException(
        "Date dropdown did not finish loading. Last visible options: "
        + repr(latest[:10])
    )


def option_matches(
    text: str,
    value: str,
    *,
    wanted_text: str | None = None,
    wanted_value: str | None = None,
    wanted_date: str | None = None,
    wanted_time: str | None = None,
) -> bool:
    normalized_text = normalize_space(text)
    normalized_value = normalize_space(value)

    if wanted_value and normalized_value == normalize_space(wanted_value):
        return True
    if wanted_text and normalized_text.casefold() == normalize_space(wanted_text).casefold():
        return True
    if wanted_date and parse_date_text(normalized_text) == wanted_date:
        return True
    if wanted_date and normalized_value == wanted_date:
        return True
    if wanted_time and normalize_time_label(normalized_text) == normalize_time_label(wanted_time):
        return True
    return False


def normalize_time_label(value: str) -> str:
    text = normalize_space(value).upper()
    for pattern in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M%p"):
        try:
            return datetime.strptime(text, pattern).strftime("%H:%M")
        except ValueError:
            continue
    return text.lstrip("0") or "0"


def select_live_option(
    driver: WebDriver,
    field_name: str,
    timeout: int,
    *,
    wanted_text: str | None = None,
    wanted_value: str | None = None,
    wanted_date: str | None = None,
    wanted_time: str | None = None,
) -> tuple[str, str]:
    """Poll the current DOM and select a matching option by index.

    Re-fetching the Select and its options on every poll avoids cached/stale
    options while the site asynchronously populates its dropdowns.
    """

    last_options: list[tuple[str, str]] = []

    def select_when_present(current_driver: WebDriver):
        nonlocal last_options
        try:
            element = find_select(current_driver, field_name)
            selector = Select(element)
            last_options = []
            for index, option in enumerate(selector.options):
                text = normalize_space(option.text)
                value = normalize_space(option.get_attribute("value") or "")
                last_options.append((text, value))
                if not option.is_enabled():
                    continue
                if option_matches(
                    text,
                    value,
                    wanted_text=wanted_text,
                    wanted_value=wanted_value,
                    wanted_date=wanted_date,
                    wanted_time=wanted_time,
                ):
                    selector.select_by_index(index)
                    try:
                        current_driver.execute_script(
                            "arguments[0].dispatchEvent(new Event('input', {bubbles:true}));"
                            "arguments[0].dispatchEvent(new Event('change', {bubbles:true}));",
                            element,
                        )
                    except StaleElementReferenceException:
                        pass
                    return text, value
        except (NoSuchElementException, StaleElementReferenceException):
            return False
        return False

    try:
        return WebDriverWait(
            driver,
            timeout,
            poll_frequency=0.25,
            ignored_exceptions=(NoSuchElementException, StaleElementReferenceException),
        ).until(select_when_present)
    except TimeoutException as exc:
        criteria = {
            "text": wanted_text,
            "value": wanted_value,
            "date": wanted_date,
            "time": wanted_time,
        }
        raise TimeoutException(
            f"Could not select {field_name!r} option matching {criteria}. "
            f"Current options: {last_options}"
        ) from exc


def find_get_log_button(driver: WebDriver):
    for element in driver.find_elements(
        By.XPATH,
        "//button | //input[@type='submit'] | //input[@type='button'] | //a",
    ):
        try:
            if not element.is_displayed() or not element.is_enabled():
                continue
            label = normalize_space(
                " ".join(
                    filter(
                        None,
                        (
                            element.text,
                            element.get_attribute("value"),
                            element.get_attribute("aria-label"),
                            element.get_attribute("title"),
                        ),
                    )
                )
            ).casefold()
            if "get log" in label:
                return element
        except StaleElementReferenceException:
            continue
    raise NoSuchElementException("Get Log button is not visible yet")


def wait_for_form(driver: WebDriver, timeout: int) -> None:
    WebDriverWait(driver, timeout).until(
        lambda current: current.execute_script("return document.readyState") == "complete"
    )
    WebDriverWait(driver, timeout).until(lambda current: find_select(current, "date"))


def load_form(driver: WebDriver, timeout: int) -> None:
    driver.get(START_URL)
    wait_for_form(driver, timeout)


def select_filters(driver: WebDriver, choice: DateChoice, timeout: int) -> None:
    select_live_option(driver, "date", timeout, wanted_date=choice.iso_date)
    select_live_option(driver, "time", timeout, wanted_time="0:00")
    select_live_option(driver, "agency", timeout, wanted_text="All Agencies")


def rows_owned_by(table: Tag) -> list[Tag]:
    """Return only rows whose nearest parent table is ``table``.

    The booking page nests a separate charge table inside each person's outer
    booking row. Using ``table.find_all("tr")`` without this filter mixes the
    nested charge rows into the outer booking table and concatenates values
    such as severity, statute, arrest location, and agency.
    """

    return [
        row
        for row in table.find_all("tr")
        if row.find_parent("table") is table
    ]


def table_header_row(table: Tag) -> tuple[int, list[str]] | None:
    rows = rows_owned_by(table)
    for row_index, row in enumerate(rows):
        cells = row.find_all(["th", "td"], recursive=False)
        if not cells:
            continue
        headers = [normalize_header(cell.get_text(" ", strip=True)) for cell in cells]
        recognized = {header for header in headers if header in HEADER_ALIASES.values()}
        if EXPECTED_HEADERS.issubset(recognized):
            return row_index, headers
    return None


def find_booking_table(soup: BeautifulSoup) -> tuple[Tag, int, list[str]] | None:
    candidates: list[tuple[int, Tag, int, list[str]]] = []
    for table in soup.find_all("table"):
        header = table_header_row(table)
        if header is None:
            continue
        row_index, headers = header
        score = sum(header in HEADER_ALIASES.values() for header in headers)
        candidates.append((score, table, row_index, headers))
    if not candidates:
        return None

    _, table, row_index, headers = max(
        candidates,
        key=lambda item: (item[0], -len(item[1].find_all("table"))),
    )
    return table, row_index, headers


def cell_text(cell: Tag) -> str:
    lines = [normalize_space(part) for part in cell.get_text("\n", strip=True).splitlines()]
    return "\n".join(part for part in lines if part)


def cell_text_without_nested_tables(cell: Tag) -> str:
    """Read a cell without flattening any nested charge table into its text."""

    fragment = BeautifulSoup(str(cell), "html.parser")
    clone = fragment.find(cell.name)
    if clone is None:
        return ""
    for nested in clone.find_all("table"):
        nested.decompose()
    return cell_text(clone)


def positive_int(value: object, default: int = 1) -> int:
    try:
        parsed = int(str(value))
        return parsed if parsed > 0 else default
    except (TypeError, ValueError):
        return default


def expand_html_table(table: Tag) -> list[list[str]]:
    """Return a rectangular table while honoring rowspan and colspan.

    Only rows owned by this exact table are included. Nested charge-table rows
    are parsed separately by ``raw_booking_rows``.
    """

    expanded: list[list[str]] = []
    active: dict[int, list[object]] = {}

    for row in rows_owned_by(table):
        direct = row.find_all(["th", "td"], recursive=False)
        if not direct:
            continue

        values: dict[int, str] = {}
        for column, span in list(active.items()):
            values[column] = str(span[1])
            span[0] = int(span[0]) - 1
            if int(span[0]) <= 0:
                del active[column]

        column = 0
        for cell in direct:
            while column in values:
                column += 1

            value = cell_text_without_nested_tables(cell)
            rowspan = positive_int(cell.get("rowspan"), 1)
            colspan = positive_int(cell.get("colspan"), 1)

            for offset in range(colspan):
                target = column + offset
                while target in values:
                    target += 1
                values[target] = value
                if rowspan > 1:
                    active[target] = [rowspan - 1, value]
            column = max(values) + 1

        width = max(values.keys() | active.keys(), default=-1) + 1
        expanded.append([values.get(index, "") for index in range(width)])

    width = max((len(row) for row in expanded), default=0)
    return [row + [""] * (width - len(row)) for row in expanded]


def canonical_header_map(headers: Iterable[str]) -> list[str]:
    return [HEADER_ALIASES.get(normalize_header(header), normalize_header(header)) for header in headers]


def _expanded_row_cells(
    row: Tag,
    active: dict[int, list[object]],
) -> list[tuple[str, Tag | None]]:
    """Expand one outer-table row and retain the source cell for each column."""

    values: dict[int, tuple[str, Tag | None]] = {}
    for column, span in list(active.items()):
        values[column] = (str(span[1]), span[2] if isinstance(span[2], Tag) else None)
        span[0] = int(span[0]) - 1
        if int(span[0]) <= 0:
            del active[column]

    column = 0
    for cell in row.find_all(["th", "td"], recursive=False):
        while column in values:
            column += 1

        value = cell_text_without_nested_tables(cell)
        rowspan = positive_int(cell.get("rowspan"), 1)
        colspan = positive_int(cell.get("colspan"), 1)

        for offset in range(colspan):
            target = column + offset
            while target in values:
                target += 1
            values[target] = (value, cell)
            if rowspan > 1:
                active[target] = [rowspan - 1, value, cell]
        column = max(values) + 1

    width = max(values.keys() | active.keys(), default=-1) + 1
    return [values.get(index, ("", None)) for index in range(width)]


def _nested_charge_codes(cell: Tag, outer_table: Tag) -> list[str]:
    """Extract one statute code from each row of a nested charge table."""

    result: list[str] = []
    nested_tables = [
        nested
        for nested in cell.find_all("table")
        if nested.find_parent("table") is outer_table
    ]

    for nested in nested_tables:
        rows = rows_owned_by(nested)
        charge_index: int | None = None
        header_index: int | None = None

        for index, row in enumerate(rows):
            cells = row.find_all(["th", "td"], recursive=False)
            headers = canonical_header_map(
                cell_text_without_nested_tables(candidate) for candidate in cells
            )
            if "charge" in headers:
                charge_index = headers.index("charge")
                header_index = index
                break

        for index, row in enumerate(rows):
            if header_index is not None and index <= header_index:
                continue

            cells = row.find_all(["th", "td"], recursive=False)
            if not cells:
                continue
            values = [cell_text_without_nested_tables(candidate) for candidate in cells]

            raw_charge = ""
            if charge_index is not None and charge_index < len(values):
                raw_charge = values[charge_index]
            elif len(values) >= 2 and normalize_space(values[0]).upper() in {
                "F", "M", "I", "E", "W"
            }:
                # The live page commonly has no header inside the nested table:
                # F/M/I | Charge | Arrest Location | Agency
                raw_charge = values[1]
            else:
                # Last-resort fallback: choose the first cell containing a
                # statute-shaped token, never the location or agency text.
                for value in values:
                    if EMBEDDED_CHARGE_SECTION_RE.search(value.upper().replace("§", "")):
                        raw_charge = value
                        break

            result.extend(charge_codes(raw_charge))

    return result


def raw_booking_rows(table: Tag, header_row_index: int) -> list[dict[str, str]]:
    """Return one normalized row per charge from the real booking markup.

    Each person is represented by an outer booking row. Its charges can be a
    normal set of columns or a nested four-column table. This function parses
    the nested table row-by-row so severity/location/agency can never become
    part of the charge code.
    """

    owned_rows = rows_owned_by(table)
    if header_row_index >= len(owned_rows):
        return []

    header_cells = owned_rows[header_row_index].find_all(["th", "td"], recursive=False)
    headers = canonical_header_map(
        cell_text_without_nested_tables(cell) for cell in header_cells
    )

    records: list[dict[str, str]] = []
    active: dict[int, list[object]] = {}
    previous_identity: dict[str, str] = {}
    identity_fields = (
        "booking_number",
        "id_number",
        "name",
        "sex",
        "dob",
        "date_time",
    )

    for row in owned_rows[header_row_index + 1 :]:
        slots = _expanded_row_cells(row, active)
        if len(slots) < len(headers):
            slots.extend([("", None)] * (len(headers) - len(slots)))

        record: dict[str, str] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = slots[index][0] if index < len(slots) else ""
            record[header] = normalize_space(value.replace("\n", " "))

        page_row_text = normalize_space(row.get_text(" ", strip=True)).casefold()
        if "no bookings found" in page_row_text:
            continue

        for field in identity_fields:
            if not record.get(field) and previous_identity.get(field):
                record[field] = previous_identity[field]

        if any(record.get(field) for field in identity_fields):
            for field in identity_fields:
                if record.get(field):
                    previous_identity[field] = record[field]

        nested_codes: list[str] = []
        seen_cells: set[int] = set()
        for _, source_cell in slots:
            if source_cell is None or id(source_cell) in seen_cells:
                continue
            seen_cells.add(id(source_cell))
            nested_codes.extend(_nested_charge_codes(source_cell, table))

        codes = nested_codes or charge_codes(record.get("charge", ""))
        if not codes:
            continue

        for code in codes:
            charge_record = dict(record)
            charge_record["charge"] = code
            records.append(charge_record)

    return records

def split_person_name(value: str) -> tuple[str, str, str]:
    text = normalize_space(value)
    if not text:
        return "", "", ""

    if "," in text:
        lastname, given = [normalize_space(part) for part in text.split(",", 1)]
    else:
        parts = text.split()
        if len(parts) == 1:
            return parts[0], "", ""
        lastname, given = parts[-1], " ".join(parts[:-1])

    given_parts = given.split()
    firstname = given_parts[0] if given_parts else ""
    middle = " ".join(given_parts[1:])
    return lastname, firstname, middle


def normalize_date(value: str) -> str:
    text = normalize_space(value)
    for pattern in (
        "%m/%d/%Y",
        "%m-%d-%Y",
        "%Y-%m-%d",
        "%B %d, %Y",
        "%b %d, %Y",
    ):
        try:
            return datetime.strptime(text, pattern).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return text


def normalize_datetime(value: str) -> str:
    text = normalize_space(value)
    for pattern in (
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M %p",
        "%Y-%m-%d %H:%M:%S",
        "%B %d, %Y %H:%M:%S",
        "%B %d, %Y %H:%M",
    ):
        try:
            return datetime.strptime(text, pattern).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return text


CHARGE_SECTION_RE = re.compile(
    r"(?<![A-Z0-9])"
    r"(\d{2,6}(?:\.\d+)?(?:\([A-Z0-9]+\))*)"
    r"(?![A-Z0-9])",
    re.IGNORECASE,
)

# Used for malformed/flattened cells such as:
# M2800.1(A)ARROYOGRANDEAGPDI12500(A)ARROYOGRANDEAGPD
# It deliberately extracts only numeric statute sections and ignores the
# severity letters, locations, and agency abbreviations.
EMBEDDED_CHARGE_SECTION_RE = re.compile(
    r"\d{2,6}(?:\.\d+)?(?:\([A-Z0-9]+\))*",
    re.IGNORECASE,
)


def charge_code_from_line(value: str) -> str:
    """Extract the first statute section from one visual charge line."""

    line = normalize_space(value).upper().replace("§", "")
    if not line:
        return ""

    match = EMBEDDED_CHARGE_SECTION_RE.search(line)
    if not match:
        return ""
    return normalize_charge_code(match.group(0))


def charge_codes(value: str) -> list[str]:
    """Return one charge object source value per statute found, in order.

    Normally each visual line contains exactly one code. If the HTML was
    flattened into a string, all statute-shaped values are recovered while
    text such as ``M``, ``ARROYO GRANDE``, and ``AGPD`` is ignored.
    """

    result: list[str] = []
    normalized = value.replace("\r\n", "\n").replace("\r", "\n")
    for line in normalized.split("\n"):
        cleaned = normalize_space(line).upper().replace("§", "")
        if not cleaned:
            continue
        matches = EMBEDDED_CHARGE_SECTION_RE.findall(cleaned)
        for match in matches:
            code = normalize_charge_code(match)
            if code:
                result.append(code)
    return result


STRICT_CHARGE_CODE_RE = re.compile(
    r"^\d{2,6}(?:\.\d+)?(?:\([A-Z0-9]+\))*$",
    re.IGNORECASE,
)


def explode_charge_value(value: object) -> list[str]:
    """Extract every statute code from a possibly concatenated charge value.

    The SLO page may flatten several charge-table rows into a value such as::

        M2800.1(A)ARROYOGRANDEAGPDI12500(A)ARROYOGRANDEAGPD

    This function returns exactly::

        ["2800.1(A)", "12500(A)"]

    Severity letters, arrest locations, and agency abbreviations are never
    emitted. Bare sections such as ``21310`` and ``485`` are also supported.
    Source order and duplicate charges are preserved.
    """

    raw = str(value or "").upper().replace("§", "")
    codes: list[str] = []
    for match in EMBEDDED_CHARGE_SECTION_RE.finditer(raw):
        code = normalize_charge_code(match.group(0))
        if code and STRICT_CHARGE_CODE_RE.fullmatch(code):
            codes.append(code)
    return codes


def normalize_bookings_for_output(
    bookings: Iterable[dict[str, object]],
    translator: "OffenseTranslator",
) -> list[dict[str, object]]:
    """Apply a mandatory final schema and charge normalization pass.

    This runs immediately before JSON is written. Therefore, even if an
    upstream HTML parser accidentally supplies a flattened charge value, the
    saved file still contains one person object and one ``{code, name}``
    object for each statute found in that value.
    """

    normalized_people: list[dict[str, object]] = []

    # Merge any duplicate person rows before normalizing their charges.
    for booking in consolidate_people(bookings):
        output: dict[str, object] = {
            "booking_id": normalize_space(str(booking.get("booking_id", ""))),
            "id": normalize_space(str(booking.get("id", ""))),
            "lastname": normalize_space(str(booking.get("lastname", ""))),
            "firstname": normalize_space(str(booking.get("firstname", ""))),
            "middle": normalize_space(str(booking.get("middle", ""))),
            "sex": normalize_space(str(booking.get("sex", ""))).upper(),
            "dob": normalize_space(str(booking.get("dob", ""))),
            "booked_at": normalize_space(str(booking.get("booked_at", ""))),
            "charges": [],
        }

        output_charges = output["charges"]
        assert isinstance(output_charges, list)

        source_charges = booking.get("charges", [])
        if not isinstance(source_charges, list):
            source_charges = []

        for source_charge in source_charges:
            if isinstance(source_charge, dict):
                raw_code = source_charge.get("code", "")
                old_name = normalize_space(str(source_charge.get("name", "")))
            else:
                raw_code = source_charge
                old_name = ""

            extracted = explode_charge_value(raw_code)
            if not extracted:
                logging.warning(
                    "Discarding non-statute charge value for booking %s: %r",
                    output["booking_id"],
                    raw_code,
                )
                continue

            for code in extracted:
                translated = translator.translate(code)
                name = normalize_space(str(translated.get("name", "")))
                if not name or name == "unknown offense":
                    # Retain an already useful name only for a single clean code.
                    if len(extracted) == 1 and old_name and old_name != "unknown offense":
                        name = old_name
                    else:
                        name = "unknown offense"
                output_charges.append({"code": code, "name": name})

        normalized_people.append(output)

    validate_booking_output(normalized_people)
    return normalized_people


def validate_booking_output(bookings: object) -> None:
    """Refuse to write malformed or concatenated charge codes."""

    if not isinstance(bookings, list):
        raise TypeError("Booking-log output must be a JSON list")

    expected_person_keys = {
        "booking_id", "id", "lastname", "firstname", "middle",
        "sex", "dob", "booked_at", "charges",
    }
    expected_charge_keys = {"code", "name"}

    for person_index, person in enumerate(bookings):
        if not isinstance(person, dict):
            raise TypeError(f"Booking entry {person_index} is not an object")
        if set(person) != expected_person_keys:
            raise ValueError(
                f"Booking entry {person_index} has wrong keys: {sorted(person)}"
            )
        charges = person.get("charges")
        if not isinstance(charges, list):
            raise TypeError(f"Booking entry {person_index} charges is not a list")
        for charge_index, charge in enumerate(charges):
            if not isinstance(charge, dict) or set(charge) != expected_charge_keys:
                raise ValueError(
                    f"Booking {person_index} charge {charge_index} has wrong shape: {charge!r}"
                )
            code = str(charge.get("code", ""))
            if not STRICT_CHARGE_CODE_RE.fullmatch(code):
                raise ValueError(
                    "Refusing to write malformed charge code "
                    f"at booking {person_index}, charge {charge_index}: {code!r}"
                )

def person_identity_key(
    *,
    person_id: str,
    booking_id: str,
    lastname: str,
    firstname: str,
    middle: str,
    dob: str,
) -> tuple[str, ...]:
    """Return a stable key that represents one person within a daily log.

    The jail's ``id`` is the preferred identity because every charge row for
    the same person should carry the same person ID.  Booking ID is only a
    fallback.  This deliberately does *not* include the charge or booking
    timestamp, because doing so would create one JSON object per charge row.
    """

    if person_id:
        return ("person_id", person_id.casefold())

    if booking_id:
        return ("booking_id", booking_id.casefold())

    name_key = "|".join(
        normalize_space(value).casefold()
        for value in (lastname, firstname, middle)
    )
    if name_key.strip("|") or dob:
        return ("name_dob", name_key, normalize_space(dob))

    # This is intentionally very unlikely to collide, while still allowing
    # continuation charge rows with the same visible identity to merge.
    return ("anonymous", name_key, normalize_space(dob))


def append_charge(
    booking: dict[str, object],
    translated: dict[str, str],
) -> None:
    """Append one source charge line as one JSON charge object."""

    code = normalize_charge_code(translated.get("code", ""))
    if not code:
        return

    charges = booking.setdefault("charges", [])
    if not isinstance(charges, list):
        raise TypeError("booking['charges'] must be a list")

    charges.append(
        {
            "code": code,
            "name": normalize_space(translated.get("name", ""))
            or "unknown offense",
        }
    )


def merge_person_record(
    target: dict[str, object],
    source: dict[str, object],
) -> None:
    """Merge another row/record into one person object."""

    scalar_fields = (
        "booking_id",
        "id",
        "lastname",
        "firstname",
        "middle",
        "sex",
        "dob",
        "booked_at",
    )

    for field in scalar_fields:
        current = normalize_space(str(target.get(field, "")))
        incoming = normalize_space(str(source.get(field, "")))
        if not current and incoming:
            target[field] = incoming

    source_charges = source.get("charges", [])
    if isinstance(source_charges, list):
        for charge in source_charges:
            if isinstance(charge, dict):
                append_charge(
                    target,
                    {
                        "code": str(charge.get("code", "")),
                        "name": str(charge.get("name", "")),
                    },
                )


def consolidate_people(
    records: Iterable[dict[str, object]],
) -> list[dict[str, object]]:
    """Guarantee one output object per person and combine all charges.

    Records are matched using any stable identifier available: person ID
    first, then booking ID, then normalized name plus DOB.  Every discovered
    identifier is registered as an alias for the same output object, so a
    continuation row that omits one field still merges correctly.
    """

    people: list[dict[str, object]] = []
    by_person_id: dict[str, dict[str, object]] = {}
    by_booking_id: dict[str, dict[str, object]] = {}
    by_name_dob: dict[tuple[str, str], dict[str, object]] = {}

    for record in records:
        lastname = normalize_space(str(record.get("lastname", "")))
        firstname = normalize_space(str(record.get("firstname", "")))
        middle = normalize_space(str(record.get("middle", "")))
        person_id = normalize_space(str(record.get("id", "")))
        booking_id = normalize_space(str(record.get("booking_id", "")))
        dob = normalize_space(str(record.get("dob", "")))

        person_key = person_id.casefold()
        booking_key = booking_id.casefold()
        name_key = "|".join(
            value.casefold() for value in (lastname, firstname, middle)
        )
        name_dob_key = (name_key, dob)

        existing = None
        if person_key:
            existing = by_person_id.get(person_key)
        if existing is None and booking_key:
            existing = by_booking_id.get(booking_key)
        if existing is None and (name_key.strip("|") or dob):
            existing = by_name_dob.get(name_dob_key)

        if existing is None:
            existing = {
                "booking_id": booking_id,
                "id": person_id,
                "lastname": lastname,
                "firstname": firstname,
                "middle": middle,
                "sex": normalize_space(str(record.get("sex", ""))).upper(),
                "dob": dob,
                "booked_at": normalize_space(str(record.get("booked_at", ""))),
                "charges": [],
            }
            people.append(existing)

        merge_person_record(existing, record)

        # Register every identifier seen on either the incoming record or the
        # merged object as an alias of this same person.
        merged_person_id = normalize_space(str(existing.get("id", ""))).casefold()
        merged_booking_id = normalize_space(
            str(existing.get("booking_id", ""))
        ).casefold()
        merged_name = "|".join(
            normalize_space(str(existing.get(field, ""))).casefold()
            for field in ("lastname", "firstname", "middle")
        )
        merged_dob = normalize_space(str(existing.get("dob", "")))

        for key in filter(None, {person_key, merged_person_id}):
            by_person_id[key] = existing
        for key in filter(None, {booking_key, merged_booking_id}):
            by_booking_id[key] = existing
        if name_key.strip("|") or dob:
            by_name_dob[name_dob_key] = existing
        if merged_name.strip("|") or merged_dob:
            by_name_dob[(merged_name, merged_dob)] = existing

    return people

def parse_booking_log(html: str, translator: OffenseTranslator) -> list[dict[str, object]]:
    soup = BeautifulSoup(html, "html.parser")
    found = find_booking_table(soup)
    if found is None:
        page_text = normalize_space(soup.get_text(" ", strip=True)).casefold()
        if "no bookings found" in page_text:
            return []
        raise RuntimeError(
            "Could not find a booking table with headers: "
            "Booking Number, ID Number, Name, Sex, DOB, Date/Time, and Charge"
        )

    table, header_row_index, _ = found
    rows = raw_booking_rows(table, header_row_index)

    # First pass: collapse all rows belonging to the same person.  Person ID is
    # primary.  This means three charge rows for ID 710979 create one booking
    # object with three entries in its charges array.
    provisional: list[dict[str, object]] = []
    by_identity: dict[tuple[str, ...], dict[str, object]] = {}

    for row in rows:
        booking_id = normalize_space(row.get("booking_number", ""))
        person_id = normalize_space(row.get("id_number", ""))
        raw_name = normalize_space(row.get("name", ""))
        lastname, firstname, middle = split_person_name(raw_name)
        dob = normalize_date(row.get("dob", ""))
        booked_at = normalize_datetime(row.get("date_time", ""))

        if not booking_id and not person_id and not raw_name:
            continue

        key = person_identity_key(
            person_id=person_id,
            booking_id=booking_id,
            lastname=lastname,
            firstname=firstname,
            middle=middle,
            dob=dob,
        )

        booking = by_identity.get(key)
        if booking is None:
            booking = {
                "booking_id": booking_id,
                "id": person_id,
                "lastname": lastname,
                "firstname": firstname,
                "middle": middle,
                "sex": normalize_space(row.get("sex", "")).upper(),
                "dob": dob,
                "booked_at": booked_at,
                "charges": [],
            }
            by_identity[key] = booking
            provisional.append(booking)
        else:
            merge_person_record(
                booking,
                {
                    "booking_id": booking_id,
                    "id": person_id,
                    "lastname": lastname,
                    "firstname": firstname,
                    "middle": middle,
                    "sex": normalize_space(row.get("sex", "")).upper(),
                    "dob": dob,
                    "booked_at": booked_at,
                    "charges": [],
                },
            )

        for code in charge_codes(row.get("charge", "")):
            append_charge(booking, translator.translate(code))

    # Defensive final pass: guarantees a single person entry even if the page
    # represented some rows differently enough to make provisional duplicates.
    return consolidate_people(provisional)

def page_contains_booking_result(html: str) -> bool:
    soup = BeautifulSoup(html, "html.parser")
    return find_booking_table(soup) is not None or "no bookings found" in normalize_space(
        soup.get_text(" ", strip=True)
    ).casefold()


def submit_and_wait(driver: WebDriver, timeout: int) -> None:
    wait = WebDriverWait(
        driver,
        timeout,
        poll_frequency=0.25,
        ignored_exceptions=(NoSuchElementException, StaleElementReferenceException),
    )
    button = wait.until(lambda current: find_get_log_button(current))
    old_html = driver.page_source

    try:
        button.click()
    except (ElementClickInterceptedException, WebDriverException):
        driver.execute_script("arguments[0].click();", button)

    def result_ready(current: WebDriver):
        try:
            return (
                current.execute_script("return document.readyState") == "complete"
                and current.page_source != old_html
                and page_contains_booking_result(current.page_source)
            )
        except WebDriverException:
            return False

    wait.until(result_ready)


def ensure_offense_workbook(
    cache_path: Path,
    *,
    refresh: bool,
    allow_download: bool,
) -> Path | None:
    if cache_path.exists() and not refresh:
        return cache_path
    if not allow_download:
        return cache_path if cache_path.exists() else None

    cache_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = cache_path.with_suffix(".xlsx")
    try:
        logging.info("Downloading California DOJ offense-code workbook")
        with requests.get(
            DOJ_WORKBOOK_URL,
            timeout=60,
            stream=True,
            headers={"User-Agent": "SLOBookingLogArchiver/2.0"},
        ) as response:
            response.raise_for_status()
            with temporary.open("wb") as output:
                for chunk in response.iter_content(chunk_size=1024 * 128):
                    if chunk:
                        output.write(chunk)
        # Validate before replacing a good cache.
        workbook = load_workbook(temporary, read_only=True, data_only=True)
        workbook.close()
        temporary.replace(cache_path)
        return cache_path
    except Exception:
        logging.exception("Could not download or validate the DOJ offense workbook")
        temporary.unlink(missing_ok=True)
        return cache_path if cache_path.exists() else None


def save_json(path: Path, data: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(data, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def process_date(
    driver: WebDriver,
    choice: DateChoice,
    translator: OffenseTranslator,
    output_dir: Path,
    timeout: int,
    overwrite: bool,
    save_html: bool,
    retries: int,
) -> Path:
    output_path = output_dir / f"{choice.iso_date}.json"
    html_path = output_dir / "html" / f"{choice.iso_date}.html"

    if output_path.exists() and not overwrite:
        logging.info("Skipping existing %s", output_path)
        return output_path

    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            load_form(driver, timeout)
            select_filters(driver, choice, timeout)
            logging.info(
                "Selected Date=%s, Time=0:00, Agency=All Agencies",
                choice.text,
            )
            submit_and_wait(driver, timeout)
            html = driver.page_source
            bookings = parse_booking_log(html, translator)
            bookings = normalize_bookings_for_output(bookings, translator)
            save_json(output_path, bookings)

            if save_html:
                html_path.parent.mkdir(parents=True, exist_ok=True)
                html_path.write_text(html, encoding="utf-8")

            logging.info("Saved %d people to %s", len(bookings), output_path)
            return output_path
        except Exception as exc:
            last_error = exc
            logging.exception(
                "Attempt %d/%d failed for %s",
                attempt,
                retries,
                choice.text,
            )
            try:
                debug_path = output_dir / "failed_html" / f"{choice.iso_date}-attempt-{attempt}.html"
                debug_path.parent.mkdir(parents=True, exist_ok=True)
                debug_path.write_text(driver.page_source, encoding="utf-8")
                logging.error("Saved failure HTML to %s", debug_path)
            except Exception:
                logging.exception("Could not save failure HTML")
            if attempt < retries:
                time.sleep(2.0 * attempt)

    assert last_error is not None
    raise last_error


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Download every available SLO Sheriff booking-log date as normalized JSON."
    )
    parser.add_argument("--output-dir", type=Path, default=Path("booking_logs"))
    parser.add_argument("--timeout", type=int, default=45)
    parser.add_argument("--delay", type=float, default=1.0)
    parser.add_argument("--retries", type=int, default=3)
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument(
        "--repair-existing",
        action="store_true",
        help="Normalize and repair existing YYYY-MM-DD.json files without crawling",
    )
    parser.add_argument("--show-browser", action="store_true")
    parser.add_argument(
        "--save-html",
        action="store_true",
        help="Save every successful result page under booking_logs/html",
    )
    parser.add_argument(
        "--refresh-offense-codes",
        action="store_true",
        help="Redownload the California DOJ offense-code workbook",
    )
    parser.add_argument(
        "--no-offense-download",
        action="store_true",
        help="Do not download the DOJ workbook; use cache and built-in names only",
    )
    return parser.parse_args()



def repair_existing_files(
    output_dir: Path,
    translator: OffenseTranslator,
) -> int:
    """Repair previously generated bare-list JSON files in place."""

    repaired = 0
    for path in sorted(output_dir.glob("????-??-??.json")):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if not isinstance(data, list):
                logging.warning("Skipping non-list legacy file %s", path)
                continue
            normalized = normalize_bookings_for_output(data, translator)
            save_json(path, normalized)
            logging.info("Repaired %s (%d people)", path, len(normalized))
            repaired += 1
        except Exception:
            logging.exception("Could not repair %s", path)
    return repaired

def main() -> int:
    args = parse_args()
    if args.retries < 1:
        raise SystemExit("--retries must be at least 1")

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )
    args.output_dir.mkdir(parents=True, exist_ok=True)

    workbook_path = ensure_offense_workbook(
        args.output_dir / ".cache" / "CIBRS_Offense_Listing.xlsx",
        refresh=args.refresh_offense_codes,
        allow_download=not args.no_offense_download,
    )
    translator = OffenseTranslator(workbook_path)
    logging.info("Running jaildriver parser version %s", SCRIPT_VERSION)
    if args.repair_existing:
        repaired = repair_existing_files(args.output_dir, translator)
        logging.info("Repaired %d existing JSON files", repaired)
        return 0
    if workbook_path is None:
        logging.warning(
            "No DOJ workbook is available. Built-in common-code names will be used; "
            "other charges will be named 'unknown offense'."
        )

    driver = build_driver(headless=not args.show_browser)
    failures: list[dict[str, str]] = []

    try:
        load_form(driver, args.timeout)
        dates = wait_for_stable_dates(driver, args.timeout)
        logging.info("Found %d dates", len(dates))

        for index, choice in enumerate(dates, start=1):
            logging.info("[%d/%d] Processing %s", index, len(dates), choice.text)
            try:
                process_date(
                    driver=driver,
                    choice=choice,
                    translator=translator,
                    output_dir=args.output_dir,
                    timeout=args.timeout,
                    overwrite=args.overwrite,
                    save_html=args.save_html,
                    retries=args.retries,
                )
            except Exception as exc:
                failures.append({"date": choice.text, "error": str(exc)})
            time.sleep(args.delay)
    finally:
        driver.quit()

    failure_path = args.output_dir / "failures.json"
    if failures:
        failure_path.write_text(
            json.dumps(failures, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
        logging.error("%d dates failed; see %s", len(failures), failure_path)
        return 1

    failure_path.unlink(missing_ok=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
